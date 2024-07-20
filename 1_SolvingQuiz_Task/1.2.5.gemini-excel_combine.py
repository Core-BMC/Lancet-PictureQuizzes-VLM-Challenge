import os
import pandas as pd
from openpyxl import Workbook, load_workbook
import json
import re


class ExcelProcessor:
    def __init__(self, base_models, variations, max_num=3):
        self.base_models = base_models
        self.variations = variations
        self.temperatures = [0, 0.5, 1]
        self.tries = range(1, 6)
        self.max_num = max_num
        self.folder_paths = self.generate_folder_paths()

    def generate_folder_paths(self):
        folder_paths = []
        for model in self.base_models:
            for variation in self.variations:
                for temp in self.temperatures:
                    for attempt in self.tries:
                        path = (f'{model}_{variation}/{model}_'
                                f'{variation}_temp_{temp}_try{attempt}')
                        folder_paths.append(path)
        return folder_paths

    @staticmethod
    def extract_info_from_text(text):
        pattern = (r'\{[^{}]*"answer":\s*"([^"]*)"[^{}]*"reason":\s*"'
                   r'([^"]*"[^{}]*)\}')
        match = re.search(pattern, text)
        if match:
            return {'answer': match.group(1), 'reason': match.group(2)}
        return None

    def process_folders(self, combined_excel_file_path):
        new_wb = Workbook()
        new_ws = new_wb.active
        new_wb.remove(new_ws)

        for folder_path in self.folder_paths:
            folder_path = folder_path.replace('.', '_')
            wb = Workbook()
            ws = wb.active
            ws.append(['case_number', 'answer', 'reason'])

            for i in range(1, self.max_num + 1):
                file_path = os.path.join(folder_path, f'{i}.txt')
                try:
                    with open(file_path, 'r', encoding='utf-8') as file:
                        content = file.read()
                        extracted_data = self.extract_info_from_text(content)
                        if extracted_data:
                            ws.append([i, extracted_data['answer'],
                                       extracted_data['reason']])
                        else:
                            ws.append([i, '', ''])
                except FileNotFoundError:
                    ws.append([i, '', ''])
                    print(f'File not found: {file_path}')
                except json.JSONDecodeError:
                    ws.append([i, '', ''])
                    print(f'Invalid JSON format: {file_path}')

            excel_file_path = os.path.join(folder_path, 'sum.xlsx')
            wb.save(excel_file_path)
            print(f'Excel file saved: {excel_file_path}')
            self.combine_sheets(new_wb, folder_path)

        new_wb.save(combined_excel_file_path)
        print(f'All content combined and saved to: {combined_excel_file_path}')

    @staticmethod
    def combine_sheets(new_wb, folder_path):
        def create_shortened_name(folder_name):
            parts = folder_name.split('_')
            if len(parts) > 4:
                return '_'.join(parts[:2] + parts[-2:])
            return folder_name

        modified_folder_path = folder_path.replace('.', '_')
        sum_file_path = os.path.join(modified_folder_path, 'sum.xlsx')
        try:
            wb = load_workbook(sum_file_path)
            ws = wb.active
            folder_name = folder_path.split('/')[-1]
            sheet_name = create_shortened_name(folder_name)
            new_ws = new_wb.create_sheet(title=sheet_name)
            for row in ws:
                new_ws.append([cell.value for cell in row])
        except FileNotFoundError:
            print(f'File not found: {sum_file_path}')


def main():
    # Original base models and variations
    original_processor = ExcelProcessor(
        base_models=['gemini', 'gemini_flash'],
        variations=['result']
    )
    original_processor.process_folders('original_combined_sum.xlsx')

    # Rephrased base models and variations
    rephrased_processor = ExcelProcessor(
        base_models=['gemini', 'gemini_flash'],
        variations=['rephrased_result']
    )
    rephrased_processor.process_folders('rephrased_combined_sum.xlsx')


if __name__ == "__main__":
    main()
